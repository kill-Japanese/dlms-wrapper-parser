#!/usr/bin/env python3
"""
DLMS/COSEM 数模 Excel 解析工具
从Excel中提取Class、Attribute、Method等数据模型信息

支持多种Excel格式，自动检测列布局和颜色方案。
"""

import sys
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# DLMS 单位编码表（部分常用单位）
DLMS_UNITS = {
    0: "unitless", 1: "year", 2: "month", 3: "week", 4: "day", 5: "hour",
    6: "min", 7: "second", 8: "degree", 9: "degree_celsius", 10: "currency",
    11: "meter", 12: "meter_per_second", 13: "cubic_meter", 14: "cubic_meter_per_hour",
    15: "W", 16: "VA", 17: "var", 18: "Wh", 19: "VAh", 20: "varh",
    21: "A", 22: "C", 23: "V", 24: "V_per_m", 25: "F", 26: "ohm",
    27: "ohm_meter", 28: "Wb", 29: "T", 30: "H", 31: "Hz",
    32: "active_energy", 33: "reactive_energy", 34: "apparent_energy",
    35: "voltage", 36: "current", 37: "power_factor", 38: "frequency",
    39: "active_power", 40: "reactive_power", 41: "apparent_power",
    42: "phase_angle", 43: "temperature", 44: "pressure", 45: "volume",
    46: "mass", 47: "length", 48: "time", 49: "speed", 50: "acceleration",
    51: "force", 52: "energy", 53: "power", 54: "charge", 55: "resistance",
    56: "capacitance", 57: "inductance", 58: "magnetic_flux",
    59: "magnetic_flux_density", 60: "magnetic_field_strength",
    61: "luminous_flux", 62: "luminous_intensity", 63: "illuminance",
    64: "luminance", 65: "amount_of_substance", 66: "molar_mass",
    67: "molar_volume", 68: "molar_concentration", 69: "mass_concentration",
    70: "density", 71: "specific_volume", 72: "concentration",
    73: "angle", 74: "solid_angle", 75: "frequency_deviation",
    76: "wavelength", 77: "wave_number", 78: "radiant_flux",
    79: "radiant_intensity", 80: "radiance", 81: "irradiance",
    82: "spectral_radiance", 83: "spectral_irradiance", 84: "radioactivity",
    85: "absorbed_dose", 86: "equivalent_dose", 87: "exposure", 88: "kerma",
    89: "catalytic_activity", 90: "catalytic_activity_concentration",
    253: "reserved1", 254: "reserved2", 255: "no_unit",
}


def get_fill_color(cell):
    """获取单元格填充色，返回十六进制颜色字符串（大写，不含alpha），无填充返回None"""
    if cell.fill and cell.fill.patternType and cell.fill.fgColor:
        rgb = cell.fill.fgColor.rgb
        if rgb and isinstance(rgb, str):
            if len(rgb) == 8:  # AARRGGBB
                return rgb[2:].upper()
            elif len(rgb) == 6:  # RRGGBB
                return rgb.upper()
    return None


def hex_to_rgb(hex_color):
    """将十六进制颜色转换为RGB元组"""
    if not hex_color or len(hex_color) != 6:
        return None
    try:
        return (int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16))
    except ValueError:
        return None


def get_color_category(color):
    """
    判断颜色类别
    返回: 'class', 'attribute', 'method', 'header', 'category', None
    
    颜色分类规则（基于常见的DLMS数模Excel格式）：
    - 浅橙色/桃色 (FFCC99等): Class行
    - 浅黄色 (FFFF99等): Attribute行
    - 极浅黄色 (FFFFCC等): Method行
    - 灰色/深灰色: 表头
    - 蓝色/浅蓝色: 分类标题
    """
    if not color:
        return None
    
    rgb = hex_to_rgb(color)
    if not rgb:
        return None
    r, g, b = rgb
    
    # 检查是否是无填充/黑色（openpyxl中000000表示无填充）
    if r == 0 and g == 0 and b == 0:
        return None
    
    # 灰色系（表头）
    if abs(r - g) < 15 and abs(g - b) < 15 and r > 150 and r < 220:
        return 'header'
    
    # 蓝色系（分类标题，如"Association & Security"）
    if b > r and b > g and b > 150:
        return 'category'
    if r < 150 and g > 150 and b > 200:  # 浅蓝色
        return 'category'
    
    # 白色或接近白色
    if r > 245 and g > 245 and b > 245:
        return None
    
    # 红色和绿色都很高（黄色系）
    if r >= 240 and g >= 240:
        # 判断蓝色分量来区分attribute和method
        if b <= 170:
            # 黄色较深: attribute (FFFF99 = b=153)
            return 'attribute'
        elif b <= 220:
            # 浅黄色: method (FFFFCC = b=204)
            return 'method'
        else:
            # 接近白色，可能是特殊标记
            return None
    
    # 橙色/桃色系（Class行）：红色高，绿色中等，蓝色较低
    # 如 FFCC99: R=255, G=204, B=153
    if r >= 240 and g >= 180 and g < 230 and b >= 120 and b < 180 and r > g and g > b:
        return 'class'
    
    # 深棕色/深橙色
    if r > 180 and g > 120 and g < 200 and b < 140 and r > g and g > b:
        return 'class'
    
    # 更通用的橙色检测：红-绿差在30-100之间，绿-蓝差在30-80之间
    if r - g >= 20 and r - g <= 100 and g - b >= 20 and g - b <= 100 and r > 180:
        return 'class'
    
    return None


def is_method_name(name):
    """
    根据名称判断是否是方法名
    方法名通常是动词开头
    """
    if not name:
        return False
    name_lower = name.lower()
    method_prefixes = [
        'add_', 'remove_', 'set_', 'get_', 'reset_', 'action_',
        'invoke_', 'execute_', 'import_', 'export_', 'update_',
        'create_', 'delete_', 'modify_', 'read_', 'write_',
        'generate_', 'transfer_', 'agreement_', 'initiate_',
        'verify_', 'authenticate_', 'activate_', 'deactivate_',
        'install_', 'uninstall_', 'upload_', 'download_',
        'start_', 'stop_', 'pause_', 'resume_',
        'key_', 'connect_', 'disconnect_', 'reply_', 'change_',
        'security_activate_', 'remove_',
    ]
    return any(name_lower.startswith(prefix) for prefix in method_prefixes)


def parse_scaler_unit(default_value):
    """解析scaler_unit格式: {-3, 15}"""
    if not default_value:
        return None, None
    
    # 匹配 {-3, 15} 或 {-3,15} 格式
    pattern = r'\{([-\d]+)\s*,\s*(\d+)\}'
    match = re.search(pattern, str(default_value))
    if match:
        scaler = int(match.group(1))
        unit = int(match.group(2))
        return scaler, unit
    return None, None


def get_unit_name(unit_code):
    """获取DLMS单位名称"""
    return DLMS_UNITS.get(unit_code, f"unknown_{unit_code}")


def safe_value(value):
    """安全获取单元格值"""
    if value is None:
        return ""
    return str(value).strip()


def parse_int(value):
    """安全解析整数"""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def detect_column_layout(ws, max_rows=20):
    """
    自动检测Excel的列布局
    
    返回列映射字典: {
        'id_col': 0,           # A列索引 - ID列
        'name_col': 1,         # B列索引 - 名称列
        'datatype_col': 2,     # C或D列索引 - 数据类型列
        'class_col': 3,        # D或E列索引 - Class ID列
        'version_col': 4,      # E或F列索引 - 版本号列
        'obis_col': 5,         # F或G列索引 - OBIS code/默认值列
        'pub_col': 6,          # G列索引 - Public权限
        'mgt_col': 7,          # H列索引 - Management权限
        'mnt_col': 8,          # I列索引 - Maintenance权限
    }
    """
    # 默认列映射（用户描述的格式）
    layout = {
        'id_col': 0,        # A - ID
        'name_col': 1,      # B - 名称
        'datatype_col': 2,  # C - 数据类型
        'class_col': 3,     # D - Class ID
        'version_col': 4,   # E - 版本
        'obis_col': 5,      # F - OBIS code/默认值
        'pub_col': 6,       # G - Public(16)
        'mgt_col': 7,       # H - Management(1)
        'mnt_col': 8,       # I - Maintenance(2)
    }
    
    # 扫描前几行来检测列布局
    headers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row)), 1):
        cells = list(row)
        for col_idx, cell in enumerate(cells):
            val = safe_value(cell.value).lower()
            if val:
                if col_idx not in headers:
                    headers[col_idx] = []
                headers[col_idx].append(val)
    
    # 根据表头关键词检测列
    class_col_candidates = []
    version_col_candidates = []
    datatype_col_candidates = []
    
    for col_idx, vals in headers.items():
        for val in vals:
            if 'class' in val:
                class_col_candidates.append(col_idx)
            if 'ver.' in val or 'version' in val:
                version_col_candidates.append(col_idx)
            if 'data type' in val or 'datatype' in val:
                datatype_col_candidates.append(col_idx)
    
    # 更新列映射
    if class_col_candidates:
        layout['class_col'] = class_col_candidates[0]
    if version_col_candidates:
        layout['version_col'] = version_col_candidates[0]
    if datatype_col_candidates:
        layout['datatype_col'] = datatype_col_candidates[0]
    
    # OBIS列通常在版本号列之后
    if layout['version_col'] >= 0:
        layout['obis_col'] = layout['version_col'] + 1
    
    # 权限列通常在OBIS列之后
    base = layout['obis_col'] + 1
    layout['pub_col'] = base
    layout['mgt_col'] = base + 1
    layout['mnt_col'] = base + 2
    
    return layout


def parse_excel(input_file, sheet_name=None, layout=None):
    """
    解析DLMS数模Excel文件
    
    识别策略（按优先级）：
    1. Class行：颜色为橙色/桃色，或Class ID列有数字
    2. Attribute行：颜色为黄色，有ID，在method区之前
    3. Method行：颜色为浅黄色，有ID，在method区中
    4. Method区的判定：
       - 颜色为method色（极浅黄色）
       - 或ID序列从较大值跳回1（ID变小或重新开始）
       - 或名称是动词开头且之前已有attribute
       - 或权限列只有Action权限（没有Get/Set）
    
    Args:
        input_file: Excel文件路径
        sheet_name: 工作表名称，None表示第一个工作表
        layout: 列映射字典，None表示自动检测
    
    Returns:
        解析结果字典
    """
    wb = load_workbook(input_file, data_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 自动检测列布局
    if layout is None:
        layout = detect_column_layout(ws)
    
    classes = []
    current_class = None
    in_method_section = False
    last_attr_id = -1
    last_method_id = -1
    current_category = ""  # 当前分类（如"Association & Security"）
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
        cells = list(row)
        
        # 获取各列值（根据检测到的列映射）
        def get_col(col_idx):
            if col_idx < len(cells):
                return safe_value(cells[col_idx].value)
            return ""
        
        a_val = get_col(layout['id_col'])
        b_val = get_col(layout['name_col'])
        c_val = get_col(layout['datatype_col'])
        d_val = get_col(layout['class_col'])
        e_val = get_col(layout['version_col'])
        f_val = get_col(layout['obis_col'])
        g_val = get_col(layout['pub_col'])
        h_val = get_col(layout['mgt_col'])
        i_val = get_col(layout['mnt_col'])
        
        # 获取B列（名称列）的颜色
        name_cell = cells[layout['name_col']] if layout['name_col'] < len(cells) else None
        b_color = get_fill_color(name_cell) if name_cell else None
        
        # 判断颜色类别
        color_cat = get_color_category(b_color)
        
        # 跳过空行
        if not b_val and not a_val and not d_val:
            continue
        
        # 跳过表头行
        if row_idx <= 3:
            header_keywords = [
                'object / attribute name', 'attribute name', 'object / attribute',
                '名称', 'name', 'data type', '数据类型', 'class', 'ver.',
                'obis code', 'access rights', 'access'
            ]
            if any(kw in b_val.lower() for kw in header_keywords):
                continue
            if color_cat == 'header':
                continue
        
        # 分类标题行（蓝色），记录分类名
        if color_cat == 'category' and b_val and not d_val and not a_val:
            current_category = b_val
            continue
        
        # ========== 识别Class行 ==========
        class_id = parse_int(d_val)
        a_id = parse_int(a_val)
        
        is_class_row = False
        
        # 1. 颜色为class色
        if color_cat == 'class' and b_val:
            is_class_row = True
        # 2. Class列有数字ID，且名称列有值，且ID列为空
        elif class_id is not None and not a_id and b_val:
            is_class_row = True
        # 3. B列颜色是class色且D/E列附近有数字
        elif color_cat == 'class':
            # 在附近列找class_id
            for col_offset in range(-2, 5):
                col_idx = layout['name_col'] + col_offset
                if 0 <= col_idx < len(cells):
                    val = parse_int(cells[col_idx].value)
                    if val is not None and val > 0 and val < 100:
                        class_id = val
                        is_class_row = True
                        break
        
        if is_class_row:
            version = parse_int(e_val)
            
            # 保存上一个class
            if current_class:
                classes.append(current_class)
            
            # 创建新class
            current_class = {
                "class_id": class_id if class_id is not None else 0,
                "class_name": b_val,
                "version": version if version is not None else 0,
                "obis_code": f_val,
                "obis_name": b_val,
                "category": current_category,
                "attributes": [],
                "methods": []
            }
            in_method_section = False
            last_attr_id = -1
            last_method_id = -1
            continue
        
        # 如果没有当前class，跳过
        if not current_class:
            continue
        
        # ========== 识别Attribute/Method行 ==========
        if a_id is not None and b_val:
            item = {
                "id": a_id,
                "name": b_val,
                "data_type": c_val if c_val else "null",
                "access_rights": {
                    "public_16": g_val if g_val else "-",
                    "management_1": h_val if h_val else "-",
                    "maintenance_2": i_val if i_val else "-"
                }
            }
            
            # 特殊处理 scaler_unit / scal_unit_type
            data_type_lower = c_val.lower() if c_val else ""
            name_lower = b_val.lower() if b_val else ""
            is_scaler_unit = (
                'scaler_unit' in data_type_lower or 
                'scal_unit' in data_type_lower or
                'scaler_unit' in name_lower
            )
            
            if is_scaler_unit:
                # 在多列中查找scaler_unit格式的值
                # 可能在obis_col、datatype_col之后的列等
                scaler = None
                unit = None
                for col_offset in range(-1, 8):
                    col_idx = layout['obis_col'] + col_offset
                    if 0 <= col_idx < len(cells):
                        cell_val = safe_value(cells[col_idx].value)
                        s, u = parse_scaler_unit(cell_val)
                        if s is not None:
                            scaler = s
                            unit = u
                            break
                
                if scaler is not None:
                    item["scaler"] = scaler
                if unit is not None:
                    item["unit"] = unit
                    item["unit_name"] = get_unit_name(unit)
            
            # 判断是attribute还是method
            is_method = False
            
            # 1. 颜色判断
            if color_cat == 'method':
                is_method = True
            elif color_cat == 'attribute':
                is_method = False
            # 2. 已在method区中
            elif in_method_section:
                is_method = True
            # 3. ID跳变检测：如果当前ID比上一个attribute ID小，说明进入了method区
            elif last_attr_id >= 2 and a_id < last_attr_id:
                is_method = True
            # 4. ID跳变检测：如果ID从一个较大值跳回1，且已有多个attribute
            elif last_attr_id >= 3 and a_id == 1:
                is_method = True
            # 5. 名称是动词开头，且已有attribute
            elif last_attr_id >= 0 and is_method_name(b_val):
                is_method = True
            # 6. 权限只有Action没有Get/Set（典型的method权限）
            elif last_attr_id >= 0:
                all_rights = f"{g_val} {h_val} {i_val}".lower()
                if 'action' in all_rights and 'get' not in all_rights and 'set' not in all_rights:
                    is_method = True
            # 7. 没有数据类型（method通常没有返回值数据类型）
            elif last_attr_id >= 2 and not c_val:
                is_method = True
            
            if is_method:
                current_class["methods"].append(item)
                in_method_section = True
                if a_id > last_method_id:
                    last_method_id = a_id
            else:
                current_class["attributes"].append(item)
                if a_id > last_attr_id:
                    last_attr_id = a_id
    
    # 添加最后一个class
    if current_class:
        classes.append(current_class)
    
    wb.close()
    
    return {
        "total_classes": len(classes),
        "layout_detected": layout,
        "classes": classes
    }


def print_summary(result):
    """打印解析摘要"""
    print(f"解析完成！共找到 {result['total_classes']} 个 Class")
    
    # 按分类分组显示
    categories = {}
    for cls in result['classes']:
        cat = cls.get('category', '未分类')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(cls)
    
    print("-" * 70)
    for cat, cls_list in categories.items():
        if cat:
            print(f"\n【{cat}】")
        for cls in cls_list:
            print(f"  Class {cls['class_id']:3d}: {cls['class_name']} (v{cls['version']})")
            print(f"    OBIS: {cls['obis_code'] or 'N/A'}")
            print(f"    Attributes: {len(cls['attributes'])} 个  |  Methods: {len(cls['methods'])} 个")


def main():
    if len(sys.argv) < 2:
        print("用法: python parse_dlms_model.py <input.xlsx> [output.json] [sheet_name]")
        print()
        print("参数:")
        print("  input.xlsx   - 输入的DLMS数模Excel文件")
        print("  output.json  - (可选) 输出的JSON文件路径")
        print("  sheet_name   - (可选) 工作表名称")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
    
    print(f"正在解析: {input_file}")
    if sheet_name:
        print(f"工作表: {sheet_name}")
    print()
    
    try:
        result = parse_excel(input_file, sheet_name)
        print_summary(result)
        
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            print(f"\n结果已保存到: {output_file}")
        else:
            print("\nJSON输出:")
            print(json.dumps(result, ensure_ascii=False, indent=2))
    
    except Exception as e:
        print(f"解析出错: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
