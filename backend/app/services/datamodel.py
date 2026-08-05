"""
数据模型管理 (COSEM Data Model)

从Excel文件加载OBIS对象模型，提供搜索和匹配功能。

支持的Excel格式:
1. 标准格式 (行式):
   - 列名: class_id, obis, name, attribute_id, data_type, description, unit, scaler
   - 或者使用更友好的中文列名: 类ID, OBIS码, 名称, 属性ID, 数据类型, 描述, 单位, 倍率

2. SABESP格式 (对象分组式):
   - Sheet名: "Object model"
   - 每个对象有一个"标题行"（D列/IC列有class_id值），后续行是该对象的属性/方法
   - A列 (#): 属性/方法序号
   - B列 (Object / Attribute Name): 对象/属性名称
   - C列 (Data Type): 数据类型
   - D列 (IC / Class): 接口类ID（仅对象标题行有值）
   - E列 (Ver.): 类版本
   - F列 (OBIS code / Default Value): OBIS码（仅标题行有值）
"""
import os
import re
import logging
from typing import Optional, Dict, Tuple, List

from openpyxl import load_workbook

from app.models.datamodel import CosemObject, CosemObjectDetail, CosemAttribute, CosemMethod
from app.utils.obis_utils import normalize_obis, obis_bytes_to_str
from app.utils.hex_utils import hex_to_bytes
from app.services.dlms_model_parser_adapter import (
    is_parser_available,
    parse_object_model_sheet,
)


class DataModelManager:
    """数据模型管理器 - 单例模式"""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True

        # 内存索引
        # 主索引: (class_id, obis_tuple, attribute_id) -> CosemObject
        self._objects: Dict[Tuple[int, Tuple[int, ...], int], CosemObject] = {}

        # 辅助索引
        self._by_class: Dict[int, List[CosemObject]] = {}
        self._all_objects: List[CosemObject] = []

        # 加载状态
        self._loaded = False
        self._source_file: Optional[str] = None

    @property
    def is_loaded(self) -> bool:
        return self._loaded

    @property
    def total_objects(self) -> int:
        return len(self._all_objects)

    @property
    def source_file(self) -> Optional[str]:
        return self._source_file

    def load_excel(self, file_path: str) -> dict:
        """
        从Excel文件加载数据模型

        自动检测Excel格式（标准行式 / SABESP对象分组式）并解析。

        Args:
            file_path: Excel文件路径

        Returns:
            dict: 加载结果信息

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: Excel格式错误
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        # 清空现有数据
        self._objects.clear()
        self._by_class.clear()
        self._all_objects.clear()

        wb = load_workbook(file_path, read_only=True, data_only=True)

        # 自动检测格式
        format_type = self._detect_format(wb)

        count = 0
        if format_type == "object_model":
            # object_model 格式使用 dlms-model-parser，直接传文件路径
            object_model_sheet = self._find_object_model_sheet(wb)
            wb.close()  # 检测完就关闭，dlms-model-parser 自己打开
            count = self._load_object_model_format(file_path, object_model_sheet)

            # 如果 object_model 解析出 0 个对象，回退尝试 SABESP 和标准格式
            if count == 0:
                logger = logging.getLogger(__name__)
                logger.warning("object_model 格式解析出 0 个对象，尝试回退到 SABESP 格式")
                wb = load_workbook(file_path, read_only=True, data_only=True)
                sabesp_sheet = self._find_object_model_sheet(wb)
                if sabesp_sheet:
                    count = self._load_sabesp_format(wb, sabesp_sheet)
                if count == 0:
                    count = self._load_standard_format(wb)
                wb.close()
                if count > 0:
                    format_type = "sabesp" if sabesp_sheet else "standard"
        elif format_type == "sabesp":
            sabesp_sheet = self._find_object_model_sheet(wb)
            if sabesp_sheet:
                count = self._load_sabesp_format(wb, sabesp_sheet)
            wb.close()
        else:
            count = self._load_standard_format(wb)
            wb.close()

        if count == 0:
            # 不抛出异常，返回空结果，由调用方判断
            logger = logging.getLogger(__name__)
            logger.warning(f"从Excel中解析出 0 个对象 (format: {format_type})")
            self._loaded = True
            self._source_file = file_path
            return {
                "total_objects": 0,
                "classes": [],
                "source_file": os.path.basename(file_path),
                "format": format_type,
                "warning": "未能从Excel文件中解析出任何对象，请检查文件格式",
            }

        self._loaded = True
        self._source_file = file_path

        return {
            "total_objects": count,
            "classes": sorted(self._by_class.keys()),
            "source_file": os.path.basename(file_path),
            "format": format_type,
        }

    def _detect_format(self, wb) -> str:
        """
        自动检测Excel格式

        检测优先级（从高到低）：
        1. object_model: 存在 "object model" sheet（大小写不敏感），且 dlms-model-parser 可用
        2. sabesp: 存在 "Object model" sheet 且符合 SABESP 格式特征
        3. standard: 标准行式格式

        Args:
            wb: openpyxl Workbook对象

        Returns:
            str: "object_model" / "sabesp" / "standard"
        """
        # 1. 检查是否有 object model 相关的 sheet（大小写不敏感）
        object_model_sheet = self._find_object_model_sheet(wb)
        if object_model_sheet and is_parser_available():
            return "object_model"

        # 2. 检查是否有 Object model 相关 sheet（大小写不敏感），尝试 SABESP 格式
        sabesp_sheet = self._find_object_model_sheet(wb)
        if sabesp_sheet:
            # 自动检测 SABESP 列布局
            col_layout = self._detect_sabesp_columns(wb[sabesp_sheet])
            if col_layout is not None:
                return "sabesp"

        # 3. 检查活动sheet的表头是否符合标准格式
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip().lower() if h else "" for h in row]
            break

        col_map = self._map_columns(headers)
        if "class_id" in col_map and "obis" in col_map:
            return "standard"

        # 默认尝试标准格式
        return "standard"

    @staticmethod
    def _find_object_model_sheet(wb) -> Optional[str]:
        """
        查找名为 "object model" 的 sheet（大小写不敏感，支持空格/下划线/横杠变体）

        Args:
            wb: openpyxl Workbook 对象

        Returns:
            匹配的 sheet 名称，未找到返回 None
        """
        target_name = "object model"
        for sheet_name in wb.sheetnames:
            # 归一化：转小写，下划线和横杠替换为空格
            name_normalized = sheet_name.strip().lower().replace("_", " ").replace("-", " ")
            if name_normalized == target_name:
                return sheet_name
        return None

    @staticmethod
    def _detect_sabesp_columns(ws) -> Optional[Dict[str, int]]:
        """
        自动检测 SABESP 格式的列布局

        扫描前 20 行，找到同时包含 Class ID（数字）和 OBIS 码（A-B:C.D.E.F格式）的行，
        确定各列的位置。

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            列布局字典，包含 class_id_col, version_col, obis_col, name_col, attr_id_col, data_type_col 的索引
            检测失败返回 None
        """
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            # 查找这一行中的 class_id 和 obis 位置
            class_id_col = None
            obis_col = None
            name_col = 1  # 默认 B 列是名称

            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                # 检测 class_id：int 或数字字符串，且值在 0-100 之间（合理的 class id 范围）
                if class_id_col is None:
                    if isinstance(val, int) and 0 < val < 1000:
                        class_id_col = col_idx
                    elif isinstance(val, str) and val.strip().isdigit() and 0 < int(val.strip()) < 1000:
                        class_id_col = col_idx

                # 检测 obis：匹配 A-B:C.D.E.F 格式
                if obis_col is None and isinstance(val, str):
                    v = val.strip()
                    if re.match(r'^\d+-\d+:\d+\.\d+\.\d+\.\d+$', v):
                        obis_col = col_idx

            # 如果同时找到了 class_id 和 obis，且在不同列
            if class_id_col is not None and obis_col is not None and class_id_col != obis_col:
                # version 列通常在 class_id 和 obis 之间
                version_col = class_id_col + 1
                if version_col >= obis_col:
                    version_col = obis_col - 1

                # attr_id 列：默认 A 列（索引0）
                attr_id_col = 0
                # data_type 列：通常在名称列和 class_id 列之间
                data_type_col = 2 if class_id_col >= 3 else class_id_col - 1

                return {
                    "class_id_col": class_id_col,
                    "version_col": version_col,
                    "obis_col": obis_col,
                    "name_col": name_col,
                    "attr_id_col": attr_id_col,
                    "data_type_col": data_type_col,
                }

        return None

    def _load_standard_format(self, wb) -> int:
        """
        加载标准行式格式的Excel

        Args:
            wb: openpyxl Workbook对象

        Returns:
            int: 解析出的对象数量
        """
        ws = wb.active

        # 读取表头
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip().lower() if h else "" for h in row]
            break

        # 列映射 - 支持中英文列名
        col_map = self._map_columns(headers)

        if "class_id" not in col_map or "obis" not in col_map:
            # 缺少必要列，返回0个对象，由上层处理
            return 0

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                obj = self._parse_row(row, col_map)
                if obj is not None:
                    self._add_object(obj)
                    count += 1
            except Exception:
                # 跳过解析失败的行
                continue

        return count

    def _load_object_model_format(self, file_path: str, sheet_name: str) -> int:
        """
        使用 dlms-model-parser 解析 object model sheet

        Args:
            file_path: Excel 文件路径
            sheet_name: object model sheet 名称

        Returns:
            int: 解析出的对象数量
        """
        # 调用 dlms-model-parser
        result = parse_object_model_sheet(file_path, sheet_name)
        if result is None:
            # 解析失败，回退到 SABESP 格式解析
            logger = logging.getLogger(__name__)
            logger.warning("dlms-model-parser 解析失败，回退到 SABESP 格式")
            wb = load_workbook(file_path, read_only=True, data_only=True)
            count = self._load_sabesp_format(wb, sheet_name)
            wb.close()
            return count

        classes = result.get("classes", [])
        count = 0

        for cls in classes:
            class_id = cls.get("class_id", 0)
            class_name = cls.get("class_name", "")
            obis_code = cls.get("obis_code", "")
            version = str(cls.get("version", ""))
            category = cls.get("category", "")

            # 跳过没有 OBIS 码的 class（可能是分类标题或模板）
            if not obis_code:
                continue

            # 验证 OBIS 格式
            try:
                normalize_obis(obis_code)
            except ValueError:
                continue

            # 1. 添加对象本身（attribute_id = 0）
            obj_desc = f"Object: {class_name}"
            if category:
                obj_desc += f" (category: {category})"

            obj = CosemObject(
                class_id=class_id,
                obis=obis_code,
                name=class_name,
                data_type="",
                description=obj_desc,
                attribute_id=0,
                unit="",
                scaler=1.0,
                version=version,
            )
            self._add_object(obj)
            count += 1

            # 2. 添加属性（attribute_id > 0）
            for attr in cls.get("attributes", []):
                attr_id = attr.get("id", 0)
                attr_name = attr.get("name", "")
                data_type = attr.get("data_type", "")

                # 处理 scaler：dlms-model-parser 输出的是整数指数（如 -3），需转为浮点数（如 0.001）
                scaler = 1.0
                scaler_exp = attr.get("scaler")
                if scaler_exp is not None:
                    try:
                        scaler = 10 ** float(scaler_exp)
                    except (ValueError, TypeError):
                        pass

                # 处理单位：优先使用单位名称，其次使用单位编码
                unit = attr.get("unit_name", "")
                if not unit and attr.get("unit") is not None:
                    unit = f"unit_{attr['unit']}"

                obj = CosemObject(
                    class_id=class_id,
                    obis=obis_code,
                    name=attr_name,
                    data_type=data_type,
                    description=f"Attribute: {attr_name}",
                    attribute_id=attr_id,
                    unit=unit,
                    scaler=scaler,
                    version=version,
                )
                self._add_object(obj)
                count += 1

            # 3. 添加方法（attribute_id < 0，用负数区分）
            for method in cls.get("methods", []):
                method_id = method.get("id", 0)
                method_name = method.get("name", "")
                data_type = method.get("data_type", "")

                obj = CosemObject(
                    class_id=class_id,
                    obis=obis_code,
                    name=method_name,
                    data_type=data_type,
                    description=f"Method: {method_name} (method_id={method_id})",
                    attribute_id=-method_id,
                    unit="",
                    scaler=1.0,
                    version="",
                )
                self._add_object(obj)
                count += 1

        return count

    def _load_sabesp_format(self, wb, sheet_name: str = "Object model", col_layout: Optional[Dict] = None) -> int:
        """
        加载SABESP对象分组式格式的Excel

        SABESP格式特点:
        - Sheet名: "Object model"
        - 对象标题行: Class ID列有数字值，A列为空，OBIS列有OBIS码
        - 属性行: A列有数字序号，B列是属性名，C列是数据类型
        - 方法行: A列有数字序号，B列是方法名，C列可能有返回类型
        - 列布局自动检测（Class ID可能在D列或E列等）

        Args:
            wb: openpyxl Workbook对象
            sheet_name: sheet 名称
            col_layout: 列布局（如为None则自动检测）

        Returns:
            int: 解析出的对象数量
        """
        ws = wb[sheet_name]

        # 自动检测列布局
        if col_layout is None:
            col_layout = self._detect_sabesp_columns(ws)
        if col_layout is None:
            # 无法检测列布局，返回0
            return 0

        class_id_col = col_layout["class_id_col"]
        version_col = col_layout["version_col"]
        obis_col = col_layout["obis_col"]
        name_col = col_layout["name_col"]
        attr_id_col = col_layout["attr_id_col"]
        data_type_col = col_layout["data_type_col"]

        count = 0
        current_object = None  # 当前对象信息
        in_method_section = False  # 是否在方法部分
        last_attr_id = 0  # 上一个属性/方法ID，用于检测方法部分的开始
        attr_count = 0  # 当前对象已解析的属性数量
        method_count = 0  # 当前对象已解析的方法数量

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            try:
                def _get(col):
                    return row[col] if len(row) > col else None

                a_val = _get(attr_id_col)
                b_val = _get(name_col)
                c_val = _get(data_type_col)
                d_val = _get(class_id_col)
                e_val = _get(version_col)
                f_val = _get(obis_col)

                # 检测对象标题行
                is_header = self._is_sabesp_object_header(a_val, d_val, f_val)

                if is_header:
                    # 解析对象标题行
                    class_id = int(str(d_val).strip())
                    obis_str = str(f_val).strip()
                    object_name = str(b_val).strip() if b_val else ""

                    # 验证OBIS格式
                    try:
                        normalize_obis(obis_str)
                    except ValueError:
                        # OBIS格式无效，跳过此对象
                        current_object = None
                        continue

                    current_object = {
                        "class_id": class_id,
                        "obis": obis_str,
                        "name": object_name,
                        "version": str(e_val).strip() if e_val else "",
                    }
                    in_method_section = False
                    last_attr_id = 0
                    attr_count = 0
                    method_count = 0

                    # 对象本身也作为一条记录（attribute_id=0）
                    obj = CosemObject(
                        class_id=class_id,
                        obis=obis_str,
                        name=object_name,
                        data_type="",
                        description=f"Object: {object_name}",
                        attribute_id=0,
                        unit="",
                        scaler=1.0,
                        version=str(e_val).strip() if e_val else "",
                    )
                    self._add_object(obj)
                    count += 1

                elif current_object and a_val is not None:
                    # 属性/方法行
                    attr_id = self._parse_attribute_id(a_val)
                    if attr_id is None or attr_id <= 0:
                        continue

                    attr_name = str(b_val).strip() if b_val else ""
                    data_type = str(c_val).strip() if c_val else ""

                    # 跳过空名称的行
                    if not attr_name:
                        continue

                    # 检测是否进入方法部分
                    # 当序号从大于1的值回到1时，说明进入了方法部分
                    # COSEM标准中属性在前，方法在后，各自从1开始编号
                    if not in_method_section and attr_id == 1 and last_attr_id > 1:
                        in_method_section = True

                    if in_method_section:
                        # 方法行 - 使用负的attribute_id来区分（-1, -2, ...）
                        method_id = attr_id
                        obj = CosemObject(
                            class_id=current_object["class_id"],
                            obis=current_object["obis"],
                            name=attr_name,
                            data_type=data_type,
                            description=f"Method: {attr_name} (method_id={method_id})",
                            attribute_id=-method_id,
                            unit="",
                            scaler=1.0,
                        )
                        method_count += 1
                    else:
                        # 属性行
                        obj = CosemObject(
                            class_id=current_object["class_id"],
                            obis=current_object["obis"],
                            name=attr_name,
                            data_type=data_type,
                            description=f"Attribute: {attr_name}",
                            attribute_id=attr_id,
                            unit="",
                            scaler=1.0,
                        )
                        attr_count += 1

                    last_attr_id = attr_id
                    self._add_object(obj)
                    count += 1

            except Exception:
                # 跳过解析失败的行
                continue

        return count

    @staticmethod
    def _is_sabesp_object_header(a_val, d_val, f_val) -> bool:
        """
        判断一行是否为SABESP格式的对象标题行

        标题行特征:
        - A列为空
        - D列(IC/Class)有数字值（class_id，可以是int或数字字符串）
        - F列(OBIS code)有OBIS格式的值（A-B:C.D.E.F）

        Args:
            a_val: A列值
            d_val: D列值
            f_val: F列值

        Returns:
            bool: 是否为对象标题行
        """
        # A列为空
        if a_val is not None and str(a_val).strip() != "":
            return False

        # D列有数字值（可以是int或数字字符串）
        if d_val is None:
            return False
        if isinstance(d_val, int):
            pass  # int类型直接通过
        elif isinstance(d_val, str):
            d_str = d_val.strip()
            if not d_str.isdigit():
                return False
        else:
            return False

        # F列有OBIS格式的值
        if f_val is None:
            return False
        if not isinstance(f_val, str):
            return False
        f_str = f_val.strip()
        if not re.match(r'^\d+-\d+:\d+\.\d+\.\d+\.\d+$', f_str):
            return False

        return True

    @staticmethod
    def _parse_attribute_id(a_val) -> Optional[int]:
        """
        解析A列的属性/方法序号

        支持的格式:
        - 整数: 1, 2, 3 -> attribute_id = 1, 2, 3
        - 浮点数: 1.0, 2.1 -> 整数部分为attribute_id

        Args:
            a_val: A列值

        Returns:
            int 或 None: 解析后的属性ID
        """
        if a_val is None:
            return None

        if isinstance(a_val, int):
            return a_val

        if isinstance(a_val, float):
            return int(a_val)

        if isinstance(a_val, str):
            a_str = a_val.strip()
            if not a_str:
                return None
            # 尝试解析为整数
            try:
                return int(a_str)
            except ValueError:
                pass
            # 尝试解析为浮点数，取整数部分
            try:
                return int(float(a_str))
            except ValueError:
                return None

        return None

    def _add_object(self, obj: CosemObject):
        """
        添加对象到索引

        Args:
            obj: CosemObject对象
        """
        key = (obj.class_id, normalize_obis(obj.obis), obj.attribute_id or 0)
        self._objects[key] = obj
        self._all_objects.append(obj)

        # 按类ID索引
        if obj.class_id not in self._by_class:
            self._by_class[obj.class_id] = []
        self._by_class[obj.class_id].append(obj)

    @staticmethod
    def _map_columns(headers: List[str]) -> Dict[str, int]:
        """
        映射列名到列索引

        支持的列名（不区分大小写）:
        - class_id / 类ID / class id / class
        - obis / OBIS码 / obis code
        - name / 名称 / 描述名
        - attribute_id / 属性ID / attr id
        - data_type / 数据类型 / type
        - description / 描述 / remark / 备注
        - unit / 单位
        - scaler / 倍率 / 系数
        """
        col_map = {}

        for i, h in enumerate(headers):
            h_lower = h.lower().strip()

            if h_lower in ("class_id", "class id", "class", "类id", "类号"):
                col_map["class_id"] = i
            elif h_lower in ("obis", "obis code", "obis码", "obis代码"):
                col_map["obis"] = i
            elif h_lower in ("name", "名称", "对象名", "描述名"):
                col_map["name"] = i
            elif h_lower in ("attribute_id", "attribute id", "attr_id", "attr id", "属性id", "属性号"):
                col_map["attribute_id"] = i
            elif h_lower in ("data_type", "data type", "datatype", "数据类型", "类型"):
                col_map["data_type"] = i
            elif h_lower in ("description", "desc", "描述", "说明", "备注", "remark"):
                col_map["description"] = i
            elif h_lower in ("unit", "单位", "计量单位"):
                col_map["unit"] = i
            elif h_lower in ("scaler", "scaling", "倍率", "系数", "比例"):
                col_map["scaler"] = i

        return col_map

    @staticmethod
    def _parse_row(row: tuple, col_map: Dict[str, int]) -> Optional[CosemObject]:
        """
        解析Excel行数据

        Args:
            row: 行数据元组
            col_map: 列映射

        Returns:
            CosemObject或None（如果行无效）
        """
        def get_val(key, default=None):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return val if val is not None else default

        # class_id - 必填
        class_id_val = get_val("class_id")
        if class_id_val is None:
            return None
        try:
            class_id = int(class_id_val)
        except (ValueError, TypeError):
            return None

        # obis - 必填
        obis_val = get_val("obis")
        if obis_val is None or str(obis_val).strip() == "":
            return None
        obis_str = str(obis_val).strip()

        # 验证OBIS格式
        try:
            normalize_obis(obis_str)
        except ValueError:
            return None

        # 可选字段
        name = str(get_val("name", "")) if get_val("name") else ""
        data_type = str(get_val("data_type", "")) if get_val("data_type") else ""
        description = str(get_val("description", "")) if get_val("description") else ""
        unit = str(get_val("unit", "")) if get_val("unit") else ""

        attribute_id = None
        attr_val = get_val("attribute_id")
        if attr_val is not None:
            try:
                attribute_id = int(attr_val)
            except (ValueError, TypeError):
                pass

        scaler = 1.0
        scaler_val = get_val("scaler")
        if scaler_val is not None:
            try:
                scaler = float(scaler_val)
            except (ValueError, TypeError):
                pass

        return CosemObject(
            class_id=class_id,
            obis=obis_str,
            name=name,
            data_type=data_type,
            description=description,
            attribute_id=attribute_id,
            unit=unit,
            scaler=scaler,
        )

    def match_obis(self, class_id: int, obis_bytes: bytes, attribute_id: int = 0) -> Optional[CosemObject]:
        """
        匹配OBIS对象

        Args:
            class_id: 类ID
            obis_bytes: OBIS码（6字节）
            attribute_id: 属性ID (0表示匹配任意属性)

        Returns:
            匹配到的CosemObject，未找到返回None
        """
        if not self._loaded:
            return None

        try:
            obis_tuple = tuple(obis_bytes)
        except Exception:
            return None

        # 精确匹配
        key = (class_id, obis_tuple, attribute_id)
        if key in self._objects:
            return self._objects[key]

        # 尝试匹配 attribute_id=0 的情况（通配）
        key = (class_id, obis_tuple, 0)
        if key in self._objects:
            return self._objects[key]

        # 尝试带通配符的OBIS匹配 (255表示通配)
        for (c_id, o_tuple, a_id), obj in self._objects.items():
            if c_id == class_id and (a_id == 0 or a_id == attribute_id):
                # 检查OBIS通配匹配
                match = True
                for o1, o2 in zip(obis_tuple, o_tuple):
                    if o1 != 255 and o2 != 255 and o1 != o2:
                        match = False
                        break
                if match:
                    return obj

        return None

    def get_object_list(self, class_id: Optional[int] = None,
                        limit: int = 100, offset: int = 0) -> List[CosemObject]:
        """
        获取对象列表（所有条目，包括属性和方法）

        Args:
            class_id: 按类ID过滤（None表示不过滤）
            limit: 返回数量限制
            offset: 偏移量

        Returns:
            对象列表
        """
        if not self._loaded:
            return []

        if class_id is not None:
            objects = self._by_class.get(class_id, [])
        else:
            objects = self._all_objects

        return objects[offset:offset + limit]

    def get_object_headers(self, class_id: Optional[int] = None,
                           limit: int = 200, offset: int = 0) -> Tuple[List[CosemObject], int]:
        """
        获取对象标题行列表（仅 attribute_id=0 的对象）

        用于前端对象列表展示，只返回对象本身（不含属性和方法行）。

        Args:
            class_id: 按类ID过滤（None表示不过滤）
            limit: 返回数量限制
            offset: 偏移量

        Returns:
            (对象标题行列表, 总数) 元组
        """
        if not self._loaded:
            return [], 0

        # 筛选 attribute_id=0 的对象标题行
        if class_id is not None:
            all_objects = self._by_class.get(class_id, [])
        else:
            all_objects = self._all_objects

        headers = [obj for obj in all_objects if obj.attribute_id == 0]
        total = len(headers)

        return headers[offset:offset + limit], total

    def search(self, keyword: str, class_id: Optional[int] = None,
               limit: int = 100) -> List[CosemObject]:
        """
        搜索OBIS对象

        Args:
            keyword: 搜索关键词（匹配名称、描述、OBIS码）
            class_id: 按类ID过滤（可选）
            limit: 返回数量限制

        Returns:
            匹配的对象列表
        """
        if not self._loaded or not keyword:
            return []

        keyword_lower = keyword.lower()
        results = []

        source = self._by_class.get(class_id, []) if class_id is not None else self._all_objects

        for obj in source:
            if (keyword_lower in obj.name.lower() or
                keyword_lower in obj.obis.lower() or
                keyword_lower in obj.description.lower()):
                results.append(obj)
                if len(results) >= limit:
                    break

        return results

    def get_object_detail(self, class_id: int, obis: str) -> Optional[CosemObjectDetail]:
        """
        获取单个对象的完整详情（包含所有属性和方法）

        Args:
            class_id: 类ID
            obis: OBIS码字符串

        Returns:
            CosemObjectDetail 或 None（如果对象不存在）
        """
        if not self._loaded:
            return None

        try:
            obis_tuple = normalize_obis(obis)
        except ValueError:
            return None

        # 先找对象本身 (attribute_id=0)
        obj_key = (class_id, obis_tuple, 0)
        base_obj = self._objects.get(obj_key)
        if base_obj is None:
            # 如果没有attribute_id=0的条目，尝试找该对象的任意属性来获取基本信息
            found = None
            for (c_id, o_tuple, a_id), obj in self._objects.items():
                if c_id == class_id and o_tuple == obis_tuple and a_id > 0:
                    found = obj
                    break
            if found is None:
                return None
            base_obj = found

        # 收集属性 (attribute_id > 0)
        attributes = []
        methods = []

        for (c_id, o_tuple, a_id), obj in self._objects.items():
            if c_id == class_id and o_tuple == obis_tuple:
                if a_id > 0:
                    # 属性
                    attributes.append(CosemAttribute(
                        attribute_id=a_id,
                        name=obj.name,
                        data_type=obj.data_type,
                        description=obj.description,
                        unit=obj.unit,
                        scaler=obj.scaler,
                    ))
                elif a_id < 0:
                    # 方法（用负的attribute_id表示）
                    methods.append(CosemMethod(
                        method_id=-a_id,
                        name=obj.name,
                        data_type=obj.data_type,
                        description=obj.description,
                    ))

        # 按ID排序
        attributes.sort(key=lambda a: a.attribute_id)
        methods.sort(key=lambda m: m.method_id)

        return CosemObjectDetail(
            class_id=class_id,
            obis=obis_bytes_to_str(bytes(obis_tuple)) if isinstance(obis_tuple, tuple) else obis,
            name=base_obj.name,
            version=base_obj.version,
            description=base_obj.description,
            attributes=attributes,
            methods=methods,
        )

    def get_classes(self) -> List[int]:
        """获取所有类ID列表"""
        return sorted(self._by_class.keys())


# 全局单例实例
data_model_manager = DataModelManager()


# 便捷函数
def load_excel(file_path: str) -> dict:
    """从Excel加载数据模型"""
    return data_model_manager.load_excel(file_path)


def match_obis(class_id: int, obis_bytes: bytes, attribute_id: int = 0) -> Optional[CosemObject]:
    """匹配OBIS对象"""
    return data_model_manager.match_obis(class_id, obis_bytes, attribute_id)
